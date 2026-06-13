"""
Excel 数据库实现
"""
import os
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from .base import DatabaseBase

class ExcelDatabase(DatabaseBase):
    """Excel 数据库实现"""
    
    def __init__(self, customer_path: str, follow_up_path: str, daily_task_path: str):
        self.customer_path = customer_path
        self.follow_up_path = follow_up_path
        self.daily_task_path = daily_task_path
        self.customer_id_counter = 1
        self.follow_up_id_counter = 1
        self.daily_task_id_counter = 1
    
    def init_db(self) -> bool:
        """初始化 Excel 数据库"""
        try:
            # 创建数据目录
            os.makedirs(os.path.dirname(self.customer_path) or '.', exist_ok=True)
            
            # 初始化客户文件
            if not os.path.exists(self.customer_path):
                self._create_customer_sheet()
            
            # 初始化跟进文件
            if not os.path.exists(self.follow_up_path):
                self._create_follow_up_sheet()
            
            # 初始化每日任务文件
            if not os.path.exists(self.daily_task_path):
                self._create_daily_task_sheet()
            
            return True
        except Exception as e:
            print(f"Error initializing Excel database: {e}")
            return False
    
    def _create_customer_sheet(self):
        """创建客户信息工作表"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'customers'
        
        headers = ['ID', '姓名', '电话', '邮箱', '公司', '职位', '地址', '备注', '创建时间', '更新时间']
        ws.append(headers)
        
        # 设置样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        wb.save(self.customer_path)
    
    def _create_follow_up_sheet(self):
        """创建跟进维护工作表"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'follow_ups'
        
        headers = ['ID', '客户ID', '跟进内容', '跟进时间', '下次跟进时间', '提醒状态', '创建时间', '更新时间']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        wb.save(self.follow_up_path)
    
    def _create_daily_task_sheet(self):
        """创建每日工作任务工作表"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'daily_tasks'
        
        headers = ['ID', '任务名称', '任务日期', '优先级', '状态', '描述', '完成时间', '创建时间', '更新时间']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        wb.save(self.daily_task_path)
    
    # ==================== 客户管理 ====================
    def add_customer(self, name: str, phone: str = '', email: str = '', 
                     company: str = '', position: str = '', address: str = '', 
                     remark: str = '') -> bool:
        """添加客户"""
        try:
            wb = load_workbook(self.customer_path)
            ws = wb.active
            
            customer_id = ws.max_row
            now = datetime.now().isoformat()
            
            ws.append([customer_id, name, phone, email, company, position, address, remark, now, now])
            wb.save(self.customer_path)
            return True
        except Exception as e:
            print(f"Error adding customer: {e}")
            return False
    
    def get_customer(self, customer_id: int) -> Optional[Dict[str, Any]]:
        """获取客户信息"""
        try:
            wb = load_workbook(self.customer_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == customer_id:
                    return {
                        'id': row[0],
                        'name': row[1],
                        'phone': row[2],
                        'email': row[3],
                        'company': row[4],
                        'position': row[5],
                        'address': row[6],
                        'remark': row[7],
                        'created_at': row[8],
                        'updated_at': row[9]
                    }
            return None
        except Exception as e:
            print(f"Error getting customer: {e}")
            return None
    
    def get_all_customers(self, page: int = 1, per_page: int = 20) -> Tuple[List[Dict[str, Any]], int]:
        """获取所有客户"""
        try:
            wb = load_workbook(self.customer_path)
            ws = wb.active
            
            customers = []
            total = 0
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                if row[0] is None:
                    break
                total += 1
                
                if (idx - 1) // per_page == page - 1:
                    customers.append({
                        'id': row[0],
                        'name': row[1],
                        'phone': row[2],
                        'email': row[3],
                        'company': row[4],
                        'position': row[5],
                        'address': row[6],
                        'remark': row[7],
                        'created_at': row[8],
                        'updated_at': row[9]
                    })
            
            return customers, total
        except Exception as e:
            print(f"Error getting all customers: {e}")
            return [], 0
    
    def search_customers(self, keyword: str) -> List[Dict[str, Any]]:
        """搜索客户"""
        try:
            wb = load_workbook(self.customer_path)
            ws = wb.active
            
            customers = []
            keyword = keyword.lower()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                
                if (keyword in str(row[1]).lower() or 
                    keyword in str(row[2]).lower() or 
                    keyword in str(row[3]).lower()):
                    customers.append({
                        'id': row[0],
                        'name': row[1],
                        'phone': row[2],
                        'email': row[3],
                        'company': row[4],
                        'position': row[5],
                        'address': row[6],
                        'remark': row[7],
                        'created_at': row[8],
                        'updated_at': row[9]
                    })
            
            return customers
        except Exception as e:
            print(f"Error searching customers: {e}")
            return []
    
    def update_customer(self, customer_id: int, **kwargs) -> bool:
        """更新客户信息"""
        try:
            wb = load_workbook(self.customer_path)
            ws = wb.active
            now = datetime.now().isoformat()
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == customer_id:
                    if 'name' in kwargs:
                        row[1].value = kwargs['name']
                    if 'phone' in kwargs:
                        row[2].value = kwargs['phone']
                    if 'email' in kwargs:
                        row[3].value = kwargs['email']
                    if 'company' in kwargs:
                        row[4].value = kwargs['company']
                    if 'position' in kwargs:
                        row[5].value = kwargs['position']
                    if 'address' in kwargs:
                        row[6].value = kwargs['address']
                    if 'remark' in kwargs:
                        row[7].value = kwargs['remark']
                    row[9].value = now
                    
                    wb.save(self.customer_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error updating customer: {e}")
            return False
    
    def delete_customer(self, customer_id: int) -> bool:
        """删除客户"""
        try:
            wb = load_workbook(self.customer_path)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                if row[0].value == customer_id:
                    ws.delete_rows(idx, 1)
                    wb.save(self.customer_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error deleting customer: {e}")
            return False
    
    # ==================== 跟进维护 ====================
    def add_follow_up(self, customer_id: int, content: str, 
                      follow_up_date: datetime, 
                      next_follow_up_date: Optional[datetime] = None) -> bool:
        """添加跟进记录"""
        try:
            wb = load_workbook(self.follow_up_path)
            ws = wb.active
            
            follow_up_id = ws.max_row
            now = datetime.now().isoformat()
            
            ws.append([
                follow_up_id,
                customer_id,
                content,
                follow_up_date.isoformat() if isinstance(follow_up_date, datetime) else str(follow_up_date),
                next_follow_up_date.isoformat() if isinstance(next_follow_up_date, datetime) else str(next_follow_up_date),
                'pending',
                now,
                now
            ])
            wb.save(self.follow_up_path)
            return True
        except Exception as e:
            print(f"Error adding follow-up: {e}")
            return False
    
    def get_follow_up(self, follow_up_id: int) -> Optional[Dict[str, Any]]:
        """获取跟进记录"""
        try:
            wb = load_workbook(self.follow_up_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == follow_up_id:
                    return {
                        'id': row[0],
                        'customer_id': row[1],
                        'content': row[2],
                        'follow_up_date': row[3],
                        'next_follow_up_date': row[4],
                        'remind_status': row[5],
                        'created_at': row[6],
                        'updated_at': row[7]
                    }
            return None
        except Exception as e:
            print(f"Error getting follow-up: {e}")
            return None
    
    def get_customer_follow_ups(self, customer_id: int) -> List[Dict[str, Any]]:
        """获取客户的跟进记录"""
        try:
            wb = load_workbook(self.follow_up_path)
            ws = wb.active
            
            follow_ups = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                if row[1] == customer_id:
                    follow_ups.append({
                        'id': row[0],
                        'customer_id': row[1],
                        'content': row[2],
                        'follow_up_date': row[3],
                        'next_follow_up_date': row[4],
                        'remind_status': row[5],
                        'created_at': row[6],
                        'updated_at': row[7]
                    })
            
            return follow_ups
        except Exception as e:
            print(f"Error getting customer follow-ups: {e}")
            return []
    
    def get_pending_reminders(self) -> List[Dict[str, Any]]:
        """获取待提醒的跟进记录"""
        try:
            wb = load_workbook(self.follow_up_path)
            ws = wb.active
            
            reminders = []
            today = datetime.now().date()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                
                if row[5] == 'pending' and row[4]:  # remind_status == 'pending'
                    try:
                        next_date = datetime.fromisoformat(str(row[4])).date()
                        if next_date <= today:
                            reminders.append({
                                'id': row[0],
                                'customer_id': row[1],
                                'content': row[2],
                                'follow_up_date': row[3],
                                'next_follow_up_date': row[4],
                                'remind_status': row[5],
                                'created_at': row[6],
                                'updated_at': row[7]
                            })
                    except:
                        pass
            
            return reminders
        except Exception as e:
            print(f"Error getting pending reminders: {e}")
            return []
    
    def update_follow_up(self, follow_up_id: int, **kwargs) -> bool:
        """更新跟进记录"""
        try:
            wb = load_workbook(self.follow_up_path)
            ws = wb.active
            now = datetime.now().isoformat()
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == follow_up_id:
                    if 'content' in kwargs:
                        row[2].value = kwargs['content']
                    if 'follow_up_date' in kwargs:
                        row[3].value = kwargs['follow_up_date']
                    if 'next_follow_up_date' in kwargs:
                        row[4].value = kwargs['next_follow_up_date']
                    if 'remind_status' in kwargs:
                        row[5].value = kwargs['remind_status']
                    row[7].value = now
                    
                    wb.save(self.follow_up_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error updating follow-up: {e}")
            return False
    
    def delete_follow_up(self, follow_up_id: int) -> bool:
        """删除跟进记录"""
        try:
            wb = load_workbook(self.follow_up_path)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                if row[0].value == follow_up_id:
                    ws.delete_rows(idx, 1)
                    wb.save(self.follow_up_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error deleting follow-up: {e}")
            return False
    
    # ==================== 每日工作 ====================
    def add_daily_task(self, task_name: str, task_date: datetime, 
                       priority: str = 'medium', 
                       description: str = '') -> bool:
        """添加日常任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            
            task_id = ws.max_row
            now = datetime.now().isoformat()
            
            ws.append([
                task_id,
                task_name,
                task_date.isoformat() if isinstance(task_date, datetime) else str(task_date),
                priority,
                'pending',
                description,
                None,
                now,
                now
            ])
            wb.save(self.daily_task_path)
            return True
        except Exception as e:
            print(f"Error adding daily task: {e}")
            return False
    
    def get_daily_task(self, task_id: int) -> Optional[Dict[str, Any]]:
        """获取任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == task_id:
                    return {
                        'id': row[0],
                        'task_name': row[1],
                        'task_date': row[2],
                        'priority': row[3],
                        'status': row[4],
                        'description': row[5],
                        'completed_at': row[6],
                        'created_at': row[7],
                        'updated_at': row[8]
                    }
            return None
        except Exception as e:
            print(f"Error getting daily task: {e}")
            return None
    
    def get_daily_tasks(self, task_date: Optional[datetime] = None) -> List[Dict[str, Any]]:
        """获取指定日期的任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            
            if task_date is None:
                task_date = datetime.now()
            
            target_date = task_date.date() if isinstance(task_date, datetime) else task_date
            tasks = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                
                try:
                    row_date = datetime.fromisoformat(str(row[2])).date()
                    if row_date == target_date:
                        tasks.append({
                            'id': row[0],
                            'task_name': row[1],
                            'task_date': row[2],
                            'priority': row[3],
                            'status': row[4],
                            'description': row[5],
                            'completed_at': row[6],
                            'created_at': row[7],
                            'updated_at': row[8]
                        })
                except:
                    pass
            
            return tasks
        except Exception as e:
            print(f"Error getting daily tasks: {e}")
            return []
    
    def get_tasks_by_status(self, status: str, task_date: Optional[datetime] = None) -> List[Dict[str, Any]]:
        """按状态获取任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            
            tasks = []
            target_date = task_date.date() if isinstance(task_date, datetime) and task_date else None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                
                if row[4] == status:
                    if target_date is None:
                        tasks.append({
                            'id': row[0],
                            'task_name': row[1],
                            'task_date': row[2],
                            'priority': row[3],
                            'status': row[4],
                            'description': row[5],
                            'completed_at': row[6],
                            'created_at': row[7],
                            'updated_at': row[8]
                        })
                    else:
                        try:
                            row_date = datetime.fromisoformat(str(row[2])).date()
                            if row_date == target_date:
                                tasks.append({
                                    'id': row[0],
                                    'task_name': row[1],
                                    'task_date': row[2],
                                    'priority': row[3],
                                    'status': row[4],
                                    'description': row[5],
                                    'completed_at': row[6],
                                    'created_at': row[7],
                                    'updated_at': row[8]
                                })
                        except:
                            pass
            
            return tasks
        except Exception as e:
            print(f"Error getting tasks by status: {e}")
            return []
    
    def update_daily_task(self, task_id: int, **kwargs) -> bool:
        """更新任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            now = datetime.now().isoformat()
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == task_id:
                    if 'task_name' in kwargs:
                        row[1].value = kwargs['task_name']
                    if 'task_date' in kwargs:
                        row[2].value = kwargs['task_date']
                    if 'priority' in kwargs:
                        row[3].value = kwargs['priority']
                    if 'status' in kwargs:
                        row[4].value = kwargs['status']
                    if 'description' in kwargs:
                        row[5].value = kwargs['description']
                    row[8].value = now
                    
                    wb.save(self.daily_task_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error updating daily task: {e}")
            return False
    
    def complete_daily_task(self, task_id: int) -> bool:
        """完成任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            now = datetime.now().isoformat()
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == task_id:
                    row[4].value = 'completed'
                    row[6].value = now
                    row[8].value = now
                    
                    wb.save(self.daily_task_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error completing daily task: {e}")
            return False
    
    def delete_daily_task(self, task_id: int) -> bool:
        """删除任务"""
        try:
            wb = load_workbook(self.daily_task_path)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                if row[0].value == task_id:
                    ws.delete_rows(idx, 1)
                    wb.save(self.daily_task_path)
                    return True
            
            return False
        except Exception as e:
            print(f"Error deleting daily task: {e}")
            return False
    
    def get_task_stats(self, task_date: Optional[datetime] = None) -> Dict[str, Any]:
        """获取任务统计信息"""
        try:
            if task_date is None:
                task_date = datetime.now()
            
            tasks = self.get_daily_tasks(task_date)
            
            total = len(tasks)
            completed = len([t for t in tasks if t['status'] == 'completed'])
            pending = len([t for t in tasks if t['status'] == 'pending'])
            high_priority = len([t for t in tasks if t['priority'] == 'high'])
            
            return {
                'date': task_date.isoformat() if isinstance(task_date, datetime) else str(task_date),
                'total': total,
                'completed': completed,
                'pending': pending,
                'high_priority': high_priority,
                'completion_rate': round(completed / total * 100, 2) if total > 0 else 0
            }
        except Exception as e:
            print(f"Error getting task stats: {e}")
            return {}
